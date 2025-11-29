import os, io, csv, re, datetime
from io import BytesIO
import boto3, pg8000

s3 = boto3.client("s3")

# ---------- strict filename parsing ----------
def parse_filename_meta_strict(key):
    base = os.path.basename(key)
    name, _ext = os.path.splitext(base)
    parts = name.split("_")
    if not parts or not parts[0].isdigit() or len(parts[0]) != 4:
        raise ValueError("Bad filename: must start with 4-digit year (e.g., 2025_South_Broward_Roster.xlsx)")
    season = int(parts[0])
    roster_idx = None
    for i, p in enumerate(parts):
        if p.lower() == "roster":
            roster_idx = i
    if roster_idx is None or roster_idx <= 1:
        raise ValueError('Bad filename: must include a final "_Roster" segment')
    team = " ".join(parts[1:roster_idx]).replace("-", " ").strip()
    if not team:
        raise ValueError("Bad filename: team name missing between year and _Roster")
    if not (2000 <= season <= 2100):
        raise ValueError("Season year out of range (2000–2100)")
    return season, team

# ---------- file readers ----------
def read_rows_from_s3(bucket, key):
    obj = s3.get_object(Bucket=bucket, Key=key)
    data = obj["Body"].read()
    print(f"⬇️  Downloaded {len(data)} bytes from s3://{bucket}/{key}")

    if key.lower().endswith(".csv"):
        print("📄 Detected CSV; parsing…")
        text = data.decode("utf-8")
        rows = list(csv.DictReader(io.StringIO(text)))
        print(f"✅ Parsed {len(rows)} rows from CSV")
        return rows

    if key.lower().endswith(".xlsx"):
        print("📘 Detected Excel (.xlsx); importing openpyxl…")
        try:
            from openpyxl import load_workbook, __version__ as oxv
            print(f"✅ openpyxl imported (v{oxv}); loading workbook…")
        except ImportError:
            raise RuntimeError("openpyxl not available; attach the 'openpyxl' layer to this function")
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)
        headers = [str(h).strip() if h is not None else "" for h in next(it)]
        print(f"🧭 Header columns: {headers}")
        out = []
        for row in it:
            rec = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    rec[headers[i]] = val
            if any(v is not None and str(v).strip() != "" for v in rec.values()):
                out.append(rec)
        print(f"✅ Parsed {len(out)} rows from Excel")
        return out

    raise RuntimeError(f"Unsupported file type: {key}")

# ---------- value cleaning ----------
def strip_value(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None

def normalize_position_code(v):
    t = strip_value(v)
    return t.upper() if t else None

def safe_to_int(v):
    if v is None:
        return None
    s = str(v).strip()
    if s == "":
        return None
    try:
        return int(float(s))
    except Exception:
        return None

def normalize_jersey_text(v):
    if v is None:
        return None
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    try:
        n = int(float(s))
        return str(n)
    except Exception:
        return s if s else None

# ---------- header normalization (NEW) ----------
def _canon(k: str) -> str:
    return re.sub(r'[^a-z0-9]', '', (k or '').strip().lower())

def normalize_headers(rows):
    """
    Map new sheet:
      No -> JerseyNumber
      First Name -> FirstName
      Last Name -> LastName
      Class -> GraduationYear
      Height (in) -> Height
      Weight -> Weight
      Position 1/2/3 -> PositionID1/2/3
    Also keeps legacy-friendly inputs.
    """
    if not rows:
        return rows
    mapping = {}
    for h in rows[0].keys():
        ck = _canon(h)
        if ck in ("no", "number", "jersey", "jerseynumber", "jersey#", "uniform"):
            mapping[h] = "JerseyNumber"
        elif ck in ("firstname", "first", "first_name"):
            mapping[h] = "FirstName"
        elif ck in ("lastname", "last", "last_name", "surname"):
            mapping[h] = "LastName"
        elif ck in ("class", "gradyear", "graduationyear"):
            mapping[h] = "GraduationYear"
        elif ck in ("heightin", "height", "ht", "heightinches"):
            mapping[h] = "Height"
        elif ck in ("weight", "wt"):
            mapping[h] = "Weight"
        elif ck in ("position1", "pos1"):
            mapping[h] = "PositionID1"
        elif ck in ("position2", "pos2"):
            mapping[h] = "PositionID2"
        elif ck in ("position3", "pos3"):
            mapping[h] = "PositionID3"
        # Compatibility with old "Position" single field:
        elif ck in ("position", "pos", "positionid"):
            mapping[h] = "PositionID1"
        # else keep original
    out = []
    for r in rows:
        nr = {}
        for k, v in r.items():
            nr[mapping.get(k, k)] = v
        out.append(nr)
    print(f"🧩 Header mapping: { {k: v for k,v in mapping.items()} }")
    for i, sample in enumerate(out[:2]):
        print(f"🔍 Sample after normalize #{i+1}: {sample}")
    return out

# ---------- DB helpers ----------
def get_conn():
    return pg8000.connect(
        host=os.environ["DB_HOST"],
        database=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASS"],
        port=int(os.environ.get("DB_PORT", "5432")),
        timeout=10
    )

def get_team_id_or_fail(cur, team_name):
    cur.execute("SELECT TeamID FROM Team WHERE lower(TeamName)=lower(%s)", (team_name,))
    r = cur.fetchone()
    if not r:
        raise RuntimeError(f"Team '{team_name}' not found. Create it first.")
    return r[0]

def get_season_id_or_fail(cur, year):
    cur.execute("SELECT SeasonID FROM Season WHERE Year=%s", (year,))
    r = cur.fetchone()
    if not r:
        raise RuntimeError(f"Season {year} not found. Create it first.")
    return r[0]

def letters_only_upper(text):
    if not text:
        return ""
    return re.sub(r"[^A-Za-z]", "", str(text)).upper()

def build_player_id_base(first_name, last_name):
    last_letters  = letters_only_upper(last_name)
    first_letters = letters_only_upper(first_name)
    last_four = (last_letters + "XXXX")[:4]
    first_two = (first_letters + "XX")[:2]
    return last_four + first_two

def get_next_available_suffix(cur, base):
    cur.execute("SELECT PlayerID FROM Players WHERE PlayerID LIKE %s", (base + "__",))
    rows = cur.fetchall()
    taken = {row[0][-2:] for row in rows} if rows else set()
    for i in range(1, 100):
        nn = f"{i:02d}"
        if nn not in taken:
            return nn
    raise RuntimeError(f"No available suffix for base {base}")

def resolve_player_id(cur, first_name, last_name):
    """
    New rule without DOB:
    - If exactly one Player row exists for (FirstName, LastName), reuse it.
    - Else allocate a new ID with base + next suffix.
    """
    cur.execute("""
        SELECT PlayerID
          FROM Players
         WHERE UPPER(FirstName)=UPPER(%s) AND UPPER(LastName)=UPPER(%s)
    """, (first_name or "", last_name or ""))
    matches = cur.fetchall()
    if matches and len(matches) == 1:
        return matches[0][0]
    base = build_player_id_base(first_name, last_name)
    return base + get_next_available_suffix(cur, base)

def insert_or_update_player(cur, player_id, first_name, last_name, height_in, weight_lb, graduation_year):
    cur.execute("SELECT 1 FROM Players WHERE PlayerID=%s", (player_id,))
    exists = cur.fetchone() is not None
    if exists:
        cur.execute("""
            UPDATE Players
               SET FirstName=%s,
                   LastName=%s,
                   Height=%s,
                   Weight=%s,
                   GraduationYear=%s
             WHERE PlayerID=%s
        """, (first_name, last_name, height_in, weight_lb, graduation_year, player_id))
    else:
        cur.execute("""
            INSERT INTO Players (PlayerID, FirstName, LastName, Height, Weight, GraduationYear)
            VALUES (%s,%s,%s,%s,%s,%s)
        """, (player_id, first_name, last_name, height_in, weight_lb, graduation_year))

def valid_position(cur, pos_id):
    pos_id = (str(pos_id).strip().upper() if pos_id is not None else None)
    if not pos_id:
        return None
    cur.execute("SELECT 1 FROM Position WHERE PositionID=%s", (pos_id,))
    return pos_id if cur.fetchone() else None

def upsert_roster_three_positions(cur, team_id, season_id, player_id, jersey_number,
                                  p1, p2, p3):
    """
    Keep a single active TeamRoster row per (Team, Season, Player, EndDate=9999-12-31),
    and set PositionID1..3 + JerseyNumber.
    """
    cur.execute("""
        SELECT RosterID FROM TeamRoster
         WHERE TeamID=%s AND PlayerID=%s AND SeasonID=%s AND EndDate='9999-12-31'
    """, (team_id, player_id, season_id))
    row = cur.fetchone()
    if row:
        cur.execute("""
            UPDATE TeamRoster
               SET PositionID1=%s, PositionID2=%s, PositionID3=%s, JerseyNumber=%s
             WHERE RosterID=%s
        """, (p1, p2, p3, jersey_number, row[0]))
    else:
        cur.execute("""
            INSERT INTO TeamRoster (TeamID, PlayerID, SeasonID, PositionID1, PositionID2, PositionID3, StartDate, JerseyNumber, EndDate)
            VALUES (%s,%s,%s,%s,%s,%s,current_date,%s,'9999-12-31')
        """, (team_id, player_id, season_id, p1, p2, p3, jersey_number))

# ---------- handler ----------
def lambda_handler(event, context):
    rec = event["Records"][0]
    bucket = rec["s3"]["bucket"]["name"]
    key = rec["s3"]["object"]["key"]
    print(f"🟢 Event received for s3://{bucket}/{key}")

    if not key.startswith("roster/"):
        print("⏭️  Skipping: key not under 'roster/'")
        return {"skipped": key}

    try:
        season_year, team_name = parse_filename_meta_strict(key)
        print(f"🔎 Parsed metadata → Team: {team_name}, Season: {season_year}")
    except ValueError as e:
        print(f"❌ Filename error: {e}")
        return {"ok": False, "bucket": bucket, "key": key, "error": str(e)}

    rows = read_rows_from_s3(bucket, key)
    rows = normalize_headers(rows)
    if not rows:
        print("⚠️  No data rows found (empty file).")
        return {"ok": False, "bucket": bucket, "key": key, "error": "Empty file or no readable rows"}

    processed = 0
    skipped = 0

    print("🔌 Connecting to database…")
    conn = get_conn()
    print("✅ Database connection established")

    try:
        with conn.cursor() as cur:
            team_id = get_team_id_or_fail(cur, team_name)
            season_id = get_season_id_or_fail(cur, season_year)
            print(f"📎 Using TeamID={team_id}, SeasonID={season_id}")

            for r in rows:
                # Required names
                first_name = strip_value(r.get("FirstName"))
                last_name  = strip_value(r.get("LastName"))
                if not first_name or not last_name:
                    skipped += 1
                    continue

                jersey = normalize_jersey_text(r.get("JerseyNumber"))
                height_in = safe_to_int(r.get("Height"))            # already inches in the new sheet
                weight_lb = safe_to_int(r.get("Weight"))
                graduation_year = safe_to_int(r.get("GraduationYear"))

                # Positions (validate against Position table)
                p1 = valid_position(cur, normalize_position_code(r.get("PositionID1")))
                p2 = valid_position(cur, normalize_position_code(r.get("PositionID2")))
                p3 = valid_position(cur, normalize_position_code(r.get("PositionID3")))

                # Allocate/reuse PlayerID (no DOB now)
                player_id = r.get("PlayerID")
                if not player_id:
                    player_id = resolve_player_id(cur, first_name, last_name)

                # Upsert player, then roster (with 3 positions)
                insert_or_update_player(cur, player_id, first_name, last_name, height_in, weight_lb, graduation_year)
                upsert_roster_three_positions(cur, team_id, season_id, player_id, jersey, p1, p2, p3)
                processed += 1

        print(f"🧾 Committing transaction (processed={processed}, skipped={skipped})")
        conn.commit()
    finally:
        conn.close()
        print("🔒 DB connection closed")

    return {
        "ok": True,
        "bucket": bucket, "key": key,
        "team": team_name, "season": season_year,
        "rows_processed": processed, "rows_skipped": skipped
    }
