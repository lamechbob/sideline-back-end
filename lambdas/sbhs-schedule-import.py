import os, io, csv, re, datetime
from io import BytesIO
import boto3, pg8000

s3 = boto3.client("s3")

# =========================
# Filename parsing (STRICT)
# =========================
def parse_schedule_filename_strict(key: str) -> int:
    """
    Require: YYYY_*_Schedule.(xlsx|csv)
    Examples:
      schedule/2025_Schedule.xlsx
      schedule/2025_South_Broward_Schedule.csv
    Returns season year (int).
    """
    base = os.path.basename(key)
    name, _ext = os.path.splitext(base)
    parts = name.split("_")

    if not parts or not parts[0].isdigit() or len(parts[0]) != 4:
        raise ValueError("Bad filename: must start with 4-digit year (e.g., 2025_Schedule.xlsx)")
    year = int(parts[0])

    if not any(p.lower() == "schedule" for p in parts[1:]):
        raise ValueError('Bad filename: must include a trailing "_Schedule" segment')

    if not (2000 <= year <= 2100):
        raise ValueError("Season year out of range (2000–2100)")
    return year

# =======================
# Read CSV / Excel (.xlsx)
# =======================
def read_rows_from_s3(bucket: str, key: str):
    obj = s3.get_object(Bucket=bucket, Key=key)
    data = obj["Body"].read()
    print(f"⬇️  Downloaded {len(data)} bytes from s3://{bucket}/{key}")

    if key.lower().endswith(".csv"):
        print("📄 Detected CSV; parsing…")
        text = data.decode("utf-8")
        rows = list(csv.DictReader(io.StringIO(text)))
        print(f"✅ Parsed {len(rows)} rows from CSV")
        return rows

    if key.lower().endswith(".xlsx"):
        print("📘 Detected Excel (.xlsx); importing openpyxl…")
        try:
            from openpyxl import load_workbook, __version__ as oxv
            print(f"✅ openpyxl imported (v{oxv}); loading workbook…")
        except ImportError:
            raise RuntimeError("openpyxl not available; attach the 'openpyxl' layer to this function")
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        it = ws.iter_rows(values_only=True)

        headers = [str(h).strip() if h is not None else "" for h in next(it)]
        print(f"🧭 Header columns: {headers}")

        out = []
        for row in it:
            rec = {}
            for i, val in enumerate(row):
                if i < len(headers) and headers[i]:
                    rec[headers[i]] = val
            if any(v is not None and str(v).strip() != "" for v in rec.values()):
                out.append(rec)
        print(f"✅ Parsed {len(out)} rows from Excel")
        return out

    raise RuntimeError(f"Unsupported file type: {key}")

# =================
# Value normalizers
# =================
def strip_value(v):
    if v is None: return None
    s = str(v).strip()
    return s if s else None

def safe_to_int(v):
    if v is None: return None
    s = str(v).strip()
    if not s: return None
    try: return int(float(s))
    except: return None

def to_iso_date(v):
    if v is None: return None
    if isinstance(v, datetime.datetime): return v.date().isoformat()
    if isinstance(v, datetime.date):     return v.isoformat()
    s = str(v).strip()
    if not s: return None
    # YYYY-MM-DD
    try:
        return datetime.date.fromisoformat(s).isoformat()
    except:
        pass
    # MM/DD/YYYY
    try:
        m, d, y = s.split("/")
        return datetime.date(int(y), int(m), int(d)).isoformat()
    except:
        return None

# ==========================
# Header normalization (flex)
# ==========================
def _canon(k: str) -> str:
    return re.sub(r'[^a-z0-9]', '', (k or '').strip().lower())

def normalize_schedule_headers(rows):
    if not rows: return rows
    mapping = {}
    for h in rows[0].keys():
        ck = _canon(h)
        if ck in ("weekno","week","wk"):            mapping[h] = "Week No"
        elif ck in ("date","gamedate","game_date"): mapping[h] = "Date"
        elif ck in ("location","venue","field"):    mapping[h] = "Location"
        elif ck in ("awayteam","away","visitor"):   mapping[h] = "Away Team"
        elif ck in ("hometeam","home"):             mapping[h] = "Home Team"
        elif ck in ("awayscore","away_score"):      mapping[h] = "Away Score"
        elif ck in ("homescore","home_score"):      mapping[h] = "Home Score"
    out = []
    for r in rows:
        nr = {}
        for k, v in r.items():
            nr[mapping.get(k, k)] = v
        out.append(nr)
    print(f"🧩 Header mapping: { {k:v for k,v in mapping.items()} }")
    for i, sample in enumerate(out[:2]):
        print(f"🔍 Sample after normalize #{i+1}: {sample}")
    return out

# ============
# DB utilities
# ============
def get_conn():
    return pg8000.connect(
        host=os.environ["DB_HOST"],
        database=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASS"],
        port=int(os.environ.get("DB_PORT", "5432")),
        timeout=10
    )

def get_season_id_or_fail(cur, year: int) -> int:
    cur.execute("SELECT SeasonID FROM Season WHERE Year=%s", (year,))
    row = cur.fetchone()
    if not row:
        raise RuntimeError(f"Season {year} not found. Create it first.")
    return row[0]

def ensure_team_exists(cur, team_name: str) -> int | None:
    """
    Find TeamID by name (case-insensitive). If not found, INSERT and return new TeamID.
    (If a numeric TeamID is provided elsewhere, we do NOT create by ID.)
    """
    if not team_name: return None
    name_clean = str(team_name).strip()
    if not name_clean: return None

    # Case-insensitive exact match
    cur.execute("SELECT TeamID FROM Team WHERE lower(TeamName)=lower(%s) LIMIT 1", (name_clean,))
    row = cur.fetchone()
    if row: return row[0]

    # Fallback letters-only comparison (e.g., "St. Thomas" == "St Thomas")
    target = re.sub(r"[^A-Za-z]", "", name_clean).upper()
    cur.execute("SELECT TeamID, TeamName FROM Team")
    for tid, tname in cur.fetchall():
        if re.sub(r"[^A-Za-z]", "", (tname or "")).upper() == target:
            return tid

    # Not found → create
    cur.execute("INSERT INTO Team(TeamName) VALUES (%s) RETURNING TeamID", (name_clean,))
    new_id = cur.fetchone()[0]
    print(f"🆕 Created Team '{name_clean}' (TeamID={new_id})")
    return new_id

def resolve_team_id_from_value(cur, value):
    """
    If value is numeric -> verify exists (no creation by ID).
    If value is a name   -> ensure_team_exists (creates if missing).
    """
    if value is None: return None
    txt = str(value).strip()
    if not txt: return None
    if txt.isdigit():
        cur.execute("SELECT TeamID FROM Team WHERE TeamID=%s", (int(txt),))
        row = cur.fetchone()
        return row[0] if row else None
    return ensure_team_exists(cur, txt)

def find_existing_game(cur, season_id, game_date_iso, home_team_id, away_team_id):
    cur.execute("""
        SELECT GameID FROM Game
         WHERE SeasonID=%s AND GameDate=%s AND HomeTeamID=%s AND AwayTeamID=%s
    """, (season_id, game_date_iso, home_team_id, away_team_id))
    r = cur.fetchone()
    return r[0] if r else None

def insert_or_update_game(cur, season_id, game_date_iso, home_team_id, away_team_id,
                          week_number, location_text):
    gid = find_existing_game(cur, season_id, game_date_iso, home_team_id, away_team_id)
    if gid:
        sets, vals = [], []
        if week_number is not None: sets.append("WeekNumber=%s"); vals.append(week_number)
        if location_text is not None: sets.append("Location=%s");  vals.append(location_text)
        if sets:
            sql = f"UPDATE Game SET {', '.join(sets)} WHERE GameID=%s"
            cur.execute(sql, (*vals, gid))
        return "updated"

    cur.execute("""
        INSERT INTO Game (SeasonID, GameDate, HomeTeamID, AwayTeamID, WeekNumber, Location, HomeScore, AwayScore)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
    """, (season_id, game_date_iso, home_team_id, away_team_id, week_number, location_text, 0, 0))
    return "inserted"

# =========
# Handler
# =========
def lambda_handler(event, context):
    rec = event["Records"][0]
    bucket = rec["s3"]["bucket"]["name"]
    key = rec["s3"]["object"]["key"]
    print(f"🟢 Event received for s3://{bucket}/{key}")

    if not key.startswith("schedule/"):
        print("⏭️  Skipping: key not under 'schedule/'")
        return {"skipped": key}

    # Season from filename (STRICT)
    try:
        season_year = parse_schedule_filename_strict(key)
        print(f"🔎 Parsed season from filename: {season_year}")
    except ValueError as e:
        print(f"❌ Filename error: {e}")
        raise

    # Read & normalize
    rows = read_rows_from_s3(bucket, key)
    rows = normalize_schedule_headers(rows)
    if not rows:
        print("⚠️  No data rows found.")
        raise RuntimeError("Empty file or no readable rows")

    # DB
    print("🔌 Connecting to database…")
    conn = get_conn()
    print("✅ Database connection established")

    inserted = updated = skipped = 0

    try:
        with conn.cursor() as cur:
            # Season must exist
            season_id = get_season_id_or_fail(cur, season_year)
            print(f"📎 Using SeasonID={season_id}")

            # Pre-resolve (and create) all named teams
            unique_teams = set()
            for r in rows:
                av = strip_value(r.get("Away Team"))
                hv = strip_value(r.get("Home Team"))
                if av: unique_teams.add(av)
                if hv: unique_teams.add(hv)

            team_id_map = {}
            for name in unique_teams:
                tid = resolve_team_id_from_value(cur, name)
                if tid is None:
                    # numeric ID referenced but doesn't exist
                    raise RuntimeError(f"Team ID reference '{name}' not found. Create it or use the team name.")
                team_id_map[name] = tid

            # Process games
            for r in rows:
                game_date_iso = to_iso_date(r.get("Date"))
                away_name     = strip_value(r.get("Away Team"))
                home_name     = strip_value(r.get("Home Team"))
                if not game_date_iso or not away_name or not home_name:
                    skipped += 1
                    continue

                week_number   = safe_to_int(r.get("Week No"))
                location_text = strip_value(r.get("Location"))
                away_team_id  = team_id_map.get(away_name)
                home_team_id  = team_id_map.get(home_name)
                if away_team_id is None or home_team_id is None:
                    skipped += 1
                    continue

                result = insert_or_update_game(
                    cur, season_id, game_date_iso, home_team_id, away_team_id,
                    week_number, location_text
                )
                if result == "inserted": inserted += 1
                elif result == "updated": updated += 1
                else: skipped += 1

            print(f"🧾 Committing (inserted={inserted}, updated={updated}, skipped={skipped})")
            conn.commit()

        return {
            "ok": True,
            "bucket": bucket, "key": key,
            "season_year": season_year,
            "inserted": inserted, "updated": updated, "skipped": skipped
        }

    finally:
        try:
            conn.rollback()
        except Exception:
            pass
        conn.close()
        print("🔒 DB connection closed")
