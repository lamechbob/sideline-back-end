import os, re, datetime
from io import BytesIO
import boto3, pg8000

s3 = boto3.client("s3")

# =========================
# Filename parsing (STRICT)
# =========================
def parse_filename_game_stats(key: str):
    """
    Expect keys like:
      game-stats/42_South_Broward_GameStats.xlsx
      game-stats/42_South_Broward_Game_Stats.xlsx
    Returns (game_id:int, team_name:str)
    """
    base = os.path.basename(key)
    name, _ext = os.path.splitext(base)
    parts = name.split("_")
    if not parts:
        raise ValueError("Bad filename: empty")

    # GameID
    try:
        game_id = int(parts[0])
    except Exception:
        raise ValueError("Bad filename: first token must be integer GameID (e.g., '42_..._GameStats.xlsx')")

    # Must end with 'GameStats' OR 'Game_Stats'
    team_tokens = None
    if parts[-1].lower() == "gamestats":
        team_tokens = parts[1:-1]
    elif len(parts) >= 2 and parts[-2].lower() == "game" and parts[-1].lower() == "stats":
        team_tokens = parts[1:-2]
    else:
        raise ValueError("Bad filename: must end with '_GameStats' (or '_Game_Stats')")

    team_name = " ".join(team_tokens).replace("-", " ").strip().replace("  ", " ")
    if not team_name:
        raise ValueError("Bad filename: team name missing between GameID and _GameStats")

    return game_id, team_name

# =======================
# Read Excel (.xlsx only)
# =======================
def read_gamestats_sheet_from_s3(bucket: str, key: str):
    obj = s3.get_object(Bucket=bucket, Key=key)
    data = obj["Body"].read()
    print(f"⬇️  Downloaded {len(data)} bytes from s3://{bucket}/{key}")

    if not key.lower().endswith(".xlsx"):
        raise RuntimeError("Only .xlsx supported for Game Stats (workbook with multiple tabs)")

    print("📘 Detected Excel (.xlsx); importing openpyxl…")
    try:
        from openpyxl import load_workbook, __version__ as oxv
        print(f"✅ openpyxl imported (v{oxv}); loading workbook…")
    except ImportError:
        raise RuntimeError("openpyxl not available; attach the 'openpyxl' layer to this function")

    wb = load_workbook(BytesIO(data), read_only=True, data_only=True)

    # Find the "Game Stats" sheet (case/spacing tolerant)
    target = None
    for ws in wb.worksheets:
        nm = (ws.title or "").strip().lower().replace("  ", " ")
        if nm == "game stats" or nm.replace(" ", "") == "gamestats":
            target = ws
            break
    if target is None:
        raise RuntimeError("Sheet 'Game Stats' not found (case-insensitive).")

    it = target.iter_rows(values_only=True)
    headers_row = next(it)
    headers = [str(h).strip() if h is not None else "" for h in headers_row]
    print(f"🧭 Header columns: {headers}")

    out = []
    for row in it:
        rec = {}
        for i, val in enumerate(row):
            if i < len(headers) and headers[i]:
                rec[headers[i]] = val
        if any(v is not None and str(v).strip() != "" for v in rec.values()):
            out.append(rec)
    print(f"✅ Parsed {len(out)} rows from 'Game Stats'")
    return out

# =================
# Value normalizers
# =================
def _canon(k: str) -> str:
    return re.sub(r'[^a-z0-9]', '', (k or '').strip().lower())

def normalize_headers(rows):
    """
    Map flexible human headers to our expected keys:
      Play No | Player No | Stat Action | Stat Type | IsTD | IsSafety | Yards (A) | Yards (B) | Yards (C) | Notes
    """
    if not rows: return rows
    mapping = {}
    for h in rows[0].keys():
        ck = _canon(h)
        if ck in ("playno","play","play#"):                mapping[h] = "Play No"
        elif ck in ("playerno","player","player#"):        mapping[h] = "Player No"
        elif ck in ("stataction","action","actionname"):   mapping[h] = "Stat Action"
        elif ck in ("stattype","type"):                    mapping[h] = "Stat Type"
        elif ck in ("istd","td"):                          mapping[h] = "IsTD"
        elif ck in ("issafety","safety"):                  mapping[h] = "IsSafety"
        elif ck in ("yardsa","yards(a)","yarda","ga"):     mapping[h] = "Yards (A)"
        elif ck in ("yardsb","yards(b)","yardb","gb"):     mapping[h] = "Yards (B)"
        elif ck in ("yardsc","yards(c)","sign","gc"):      mapping[h] = "Yards (C)"
        elif ck in ("notes","remark","comments"):          mapping[h] = "Notes"
    out = []
    for r in rows:
        nr = {}
        for k, v in r.items():
            nr[mapping.get(k, k)] = v
        out.append(nr)
    print(f"🧩 Header mapping: { {k:v for k,v in mapping.items()} }")
    for i, sample in enumerate(out[:2]):
        print(f"🔍 Sample after normalize #{i+1}: {sample}")
    return out

def clean_jersey_text(val):
    if val is None: return None
    s = str(val).strip()
    if s == "": return None
    try:
        n = int(float(s))
        return str(n)
    except Exception:
        return s

def to_bool(v):
    if v is None: return False
    s = str(v).strip().lower()
    return s in ("1","true","t","yes","y","x","✓","check","checked")

def calculate_yards(row):
    """
    Sum 'Yards (A)' + 'Yards (B)' and apply sign per 'Yards (C)'.
    - Treat missing A/B as 0
    - 'Yards (C)': 'negative' → negate; anything else → positive
    - Return TEXT
    """
    def to_num(x):
        try:
            return float(x)
        except Exception:
            return 0.0
    a = to_num(row.get("Yards (A)"))
    b = to_num(row.get("Yards (B)"))
    sign = (str(row.get("Yards (C)") or "").strip().lower())
    total = a + b
    if sign.startswith("neg"):
        total = -total
    try:
        return str(int(total))
    except Exception:
        return str(int(round(total)))

def sack_weight_for(action_name: str) -> float:
    if not action_name: return 0.0
    s = str(action_name).strip().lower()
    if s == "sack": return 1.0
    if s in ("sack assist","sackassist"): return 0.5
    return 0.0

def parse_play_no(v):
    """Return integer play number (>=1) or None if invalid/blank."""
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    try:
        n = int(float(s))
        return n if n >= 1 else None
    except Exception:
        return None

# ============
# DB utilities
# ============
def get_conn():
    return pg8000.connect(
        host=os.environ["DB_HOST"],
        database=os.environ["DB_NAME"],
        user=os.environ["DB_USER"],
        password=os.environ["DB_PASS"],
        port=int(os.environ.get("DB_PORT", "5432")),
        timeout=10
    )

def get_team_id_or_fail(cur, team_name: str):
    cur.execute("SELECT TeamID FROM Team WHERE lower(TeamName)=lower(%s)", (team_name,))
    r = cur.fetchone()
    if not r:
        raise RuntimeError(f"Team '{team_name}' not found. Import roster/teams first.")
    return r[0]

def build_roster_map(cur, team_id: int):
    """
    Return dict {(team_id_str, jersey_str) -> player_id}
    Only current rows (EndDate='9999-12-31').
    """
    cur.execute("""
        SELECT PlayerID, TeamID, JerseyNumber
          FROM TeamRoster
         WHERE TeamID=%s AND EndDate='9999-12-31'
    """, (team_id,))
    out = {}
    for pid, tid, jersey in cur.fetchall():
        key = (str(tid), clean_jersey_text(jersey))
        out[key] = pid
    print(f"👕 Active roster rows cached for TeamID={team_id}: {len(out)}")
    return out

def build_stataction_map(cur):
    """Return dict {lower(ActionName): StatActionID}"""
    cur.execute("SELECT StatActionID, ActionName FROM StatAction")
    out = {}
    for sid, name in cur.fetchall():
        out[(name or "").strip().lower()] = sid
    print(f"📚 StatAction entries cached: {len(out)}")
    return out

def delete_existing_gameplays(cur, game_id: int):
    cur.execute("SELECT COUNT(*) FROM GamePlays WHERE GameID=%s", (game_id,))
    before = cur.fetchone()[0]
    cur.execute("DELETE FROM GamePlays WHERE GameID=%s", (game_id,))
    print(f"🧹 Deleted {before} existing GamePlays for GameID={game_id}")

def insert_gameplay(cur, game_id, play_no, player_id, team_id, stat_type, stat_action_id,
                    yards, is_td, is_safety, sack_weight, source_file_name, notes):
    cur.execute("""
        INSERT INTO GamePlays (
            GameID, PlayNo, PlayerID, TeamID, StatType, StatActionID,
            Yards, IsTD, IsSafety, SackWeight, SourceFileName, CreatedAt, Notes
        )
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s, now(), %s)
    """, (game_id, play_no, player_id, team_id, stat_type, stat_action_id,
          yards, is_td, is_safety, sack_weight, source_file_name, notes))

# =========
# Handler
# =========
def lambda_handler(event, context):
    # S3:ObjectCreated event
    rec = event["Records"][0]
    bucket = rec["s3"]["bucket"]["name"]
    key = rec["s3"]["object"]["key"]
    print(f"🟢 Event received for s3://{bucket}/{key}")

    if not key.startswith("game-stats/"):
        print("⏭️  Skipping: key not under 'game-stats/'")
        return {"skipped": key}

    # Parse filename for GameID + TeamName
    try:
        game_id, team_name = parse_filename_game_stats(key)
        print(f"🔎 Parsed: GameID={game_id}, Team='{team_name}'")
    except ValueError as e:
        print(f"❌ Filename error: {e}")
        raise

    # Read rows from 'Game Stats' tab
    rows = read_gamestats_sheet_from_s3(bucket, key)
    rows = normalize_headers(rows)

    # Hard filter: keep ONLY rows with a valid Play No (no evaluation of others)
    total_rows = len(rows)
    filtered = []
    for r in rows:
        pn = parse_play_no(r.get("Play No"))
        if pn is not None:
            r["_PlayNoInt"] = pn  # cache parsed int
            filtered.append(r)
    dropped = total_rows - len(filtered)
    rows = filtered
    print(f"🧹 Dropped {dropped} rows without a valid 'Play No' (kept {len(rows)} of {total_rows})")
    if not rows:
        raise RuntimeError("No valid rows with 'Play No' found in 'Game Stats'")

    # DB
    print("🔌 Connecting to database…")
    conn = get_conn()
    print("✅ Database connection established")

    inserted = skipped = 0
    try:
        with conn.cursor() as cur:
            team_id = get_team_id_or_fail(cur, team_name)
            print(f"📎 Using TeamID={team_id}")

            roster_map = build_roster_map(cur, team_id)
            action_map = build_stataction_map(cur)

            # wipe existing stats for this game
            delete_existing_gameplays(cur, game_id)

            # process & insert
            for r in rows:
                play_no = r["_PlayNoInt"]           # validated int
                jersey  = clean_jersey_text(r.get("Player No"))
                action  = (r.get("Stat Action") or "").strip()
                stat_type = (r.get("Stat Type") or "").strip() or None

                if jersey is None or not action:
                    skipped += 1
                    continue

                player_id = roster_map.get((str(team_id), jersey))
                if not player_id:
                    print(f"⚠️  Skip: no PlayerID for jersey {jersey}")
                    skipped += 1
                    continue

                action_id = action_map.get(action.strip().lower())
                if not action_id:
                    print(f"⚠️  Skip: unknown StatAction '{action}'")
                    skipped += 1
                    continue

                yards = calculate_yards(r)
                is_td = to_bool(r.get("IsTD"))
                is_safety = to_bool(r.get("IsSafety"))
                sack_weight = sack_weight_for(action)
                notes = r.get("Notes")

                try:
                    insert_gameplay(
                        cur, game_id, play_no, player_id, team_id,
                        stat_type, action_id, yards, is_td, is_safety,
                        sack_weight, key, notes
                    )
                    inserted += 1
                except Exception as e:
                    print(f"⛔ Insert failed for PlayNo={play_no}: {e}")
                    skipped += 1

            print(f"🧾 Committing (inserted={inserted}, skipped={skipped})")
            conn.commit()

        return {
            "ok": True,
            "bucket": bucket, "key": key,
            "game_id": game_id, "team": team_name, "team_id": team_id,
            "rows_total": total_rows,
            "rows_kept_with_playno": len(rows),
            "rows_dropped_no_playno": dropped,
            "inserted": inserted,
            "skipped": skipped
        }

    finally:
        try:
            conn.rollback()
        except Exception:
            pass
        conn.close()
        print("🔒 DB connection closed")
